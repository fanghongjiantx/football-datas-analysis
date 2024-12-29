import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import time
from datetime import datetime
import re
import random
from fake_useragent import UserAgent
import logging
from requests.exceptions import ConnectionError, Timeout, RequestException
from openpyxl import load_workbook


def get_goal_times(url, headers, max_retries=5, backoff_factor=0.3):
    retries = 0
    while retries < max_retries:
        try:
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, "html.parser")
            times = []
            scorebox_div = soup.find("div", class_="scorebox")
            if scorebox_div:
                event_divs = scorebox_div.find_all("div", class_="event")
                for event in event_divs:
                    event_details = event.find_all("div")
                    for detail in event_details:
                        red_card_icon = detail.find("div", class_="event_icon red_card")
                        yellow_red_card_icon = detail.find(
                            "div", class_="event_icon yellow_red_card"
                        )
                        if red_card_icon or yellow_red_card_icon:
                            continue

                        text = detail.get_text(strip=True)
                        match = re.search(r"(\d{1,2})(\+?\d{0,2})’", text)
                        if match:
                            minutes = int(match.group(1))
                            additional_time = match.group(2)
                            if additional_time:
                                additional_time = int(additional_time.lstrip("+"))
                                time_value = -(minutes + additional_time)
                            else:
                                time_value = minutes
                            times.append(time_value)
            return times
        except (ConnectionError, ConnectionResetError) as e:
            print(f"连接错误，URL: {url}. 正在重试...")
            time.sleep(backoff_factor * (2**retries))
            retries += 1
        except Timeout as e:
            print(f"请求超时，URL: {url}. 正在重试...")
            time.sleep(backoff_factor * (2**retries))
            retries += 1
        except RequestException as e:
            print(f"请求异常，URL: {url}. 错误: {e}")
            break
    return []


def get_match_report(html_content):
    # 从给定的HTML内容中解析比赛报告链接
    soup = BeautifulSoup(html_content, "html.parser")
    match_report_tds = soup.find_all(
        "td", {"class": "left", "data-stat": "match_report"}
    )

    match_reports = []
    for td in match_report_tds:
        a_tag = td.find("a")
        if a_tag and "href" in a_tag.attrs:
            match_reports.append(a_tag["href"])

    return match_reports


def get_random_user_agent():
    ua = UserAgent()
    return ua.random


def get_headers():
    headers = {
        "User-Agent": get_random_user_agent(),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept-Encoding": "gzip, deflate",
        "Connection": "keep-alive",
    }
    return headers


def get_html(link):
    # 获取HTML内容，不再保存到本地
    response = requests.get(link)
    if response.status_code == 200:
        return response.text  # 返回HTML内容


def setup_logging():
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    logfilename = f"D:\\football\\stats\\log\\footballstats_{current_time}.log"
    logging.basicConfig(
        filename=logfilename,
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )


def run(year, league_name, xlsx_path):
    setup_logging()
    try:
        wb = load_workbook(xlsx_path)
    except FileNotFoundError:
        wb = Workbook()

    sheet_name = f"{year}-{year + 1}"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # 获取比赛报告的HTML内容
    html_content = get_html(
        f"{leagues[league_name]}{year}-{year + 1}/schedule/"
    )
    match_reports = get_match_report(html_content)
    logging.info(f"找到 {len(match_reports)} 个比赛报告。")

    ws.append(["报告", "进球时间 (分钟)"])

    for i, report in enumerate(match_reports):
        try:
            time.sleep(random.uniform(3, 6))
            headers = get_headers()
            times = get_goal_times("https://fbref.com/" + report, headers=headers)
            if times:
                for t in times:
                    ws.append([report, t])
                logging.info(
                    f"处理比赛报告 {i + 1}/{len(match_reports)}: {report}"
                )
                print(f"处理比赛报告 {i + 1}/{len(match_reports)}: {report}")
                logging.info(f"进球时间已添加: {times}")
                print(f"进球时间已添加: {times}")
        except Exception as e:
            logging.error(
                f"处理比赛报告 {i + 1}: {report}. 错误: {e}"
            )
            print(f"处理比赛报告 {i + 1}: {report}. 错误: {e}")

    wb.save(xlsx_path)
    logging.info(f"数据已保存到Excel文件: {xlsx_path}")
    print(f"数据已保存到Excel文件: {xlsx_path}")


# 联赛URL字典
leagues = {
    "Premier League": "https://fbref.com/en/comps/9/",
    "Serie A": "https://fbref.com/en/comps/11/",
    "La Liga": "https://fbref.com/en/comps/12/",
    "Ligue 1": "https://fbref.com/en/comps/13/",
    "Bundesliga": "https://fbref.com/en/comps/20/",
}

def main(start_year, end_year, league_names):
    for year in range(start_year, end_year):
        for league_name in league_names:
            if league_name in leagues:  # 确保联赛在字典中
                xlsx_path = f"D:\\football\\stats\\data\\goal_stats\\{league_name}_goals_{year}-{year + 1}.xlsx"
                run(year, league_name, xlsx_path)

if __name__ == "__main__":
    selected_leagues = ["Premier League", "Serie A"]  # 自由选择获取的联赛
    main(2022, 2024, selected_leagues)  # 传入年份范围和选择的联赛名称

