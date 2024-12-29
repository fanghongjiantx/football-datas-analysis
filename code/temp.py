import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import time
from datetime import datetime
from collections import Counter
import re
import random
from fake_useragent import UserAgent
import logging
from requests.exceptions import ConnectionError, Timeout, RequestException
from openpyxl import load_workbook, Workbook


def get_goal_times(url, headers, max_retries=5, backoff_factor=0.3):
    retries = 0
    while retries < max_retries:
        try:
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, "html.parser")
            times = []
            scorebox_div = soup.find("div", class_="scorebox")
            if scorebox_div:
                event_divs = scorebox_div.find_all("div", class_="event")
                for event in event_divs:
                    event_details = event.find_all("div")
                    for detail in event_details:
                        # Check if the event has a red card icon
                        red_card_icon = detail.find("div", class_="event_icon red_card")
                        yellow_red_card_icon = detail.find(
                            "div", class_="event_icon yellow_red_card"
                        )
                        if red_card_icon or yellow_red_card_icon:
                            continue  # Skip this event as it has a red card

                        text = detail.get_text(strip=True)
                        # Modified regex to match + sign for additional time
                        match = re.search(r"(\d{1,2})(\+?\d{0,2})’", text)
                        if match:
                            # Convert the time to an integer
                            minutes = int(match.group(1))
                            additional_time = match.group(2)
                            if additional_time:
                                # Convert additional time to an integer and adjust for negative value
                                additional_time = int(additional_time.lstrip("+"))
                                time_value = -(minutes + additional_time)
                            else:
                                time_value = minutes
                            times.append(time_value)
            return times
        except (ConnectionError, ConnectionResetError) as e:
            print(f"Connection error for URL: {url}. Retrying...")
            time.sleep(backoff_factor * (2**retries))
            retries += 1
        except Timeout as e:
            print(f"Timeout error for URL: {url}. Retrying...")
            time.sleep(backoff_factor * (2**retries))
            retries += 1
        except RequestException as e:
            print(f"RequestException for URL: {url}. Error: {e}")
            break
    return []


def get_match_report(html_locate):
    # 读取HTML文件内容
    with open(html_locate, "r", encoding="utf-8") as file:
        html_content = file.read()

    # 使用BeautifulSoup解析HTML
    soup = BeautifulSoup(html_content, "html.parser")

    # 查找所有具有特定class和data-stat属性的<td>标签
    match_report_tds = soup.find_all(
        "td", {"class": "left", "data-stat": "match_report"}
    )

    # 提取每个<td>标签下<a>标签的href属性值
    match_reports = []
    for td in match_report_tds:
        a_tag = td.find("a")
        if a_tag and "href" in a_tag.attrs:
            match_reports.append(a_tag["href"])

    return match_reports


def get_random_user_agent():
    ua = UserAgent()
    return ua.random


def get_headers():
    headers = {
        "User-Agent": get_random_user_agent(),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept-Encoding": "gzip, deflate",
        "Connection": "keep-alive",
        # 'Referer': 'https://www.example.com/',  # 如果需要，可以添加Referer头
    }
    return headers


def get_html(link, save_path):
    # 发起请求
    response = requests.get(link)

    # 检查请求是否成功
    if response.status_code == 200:
        # 解析HTML
        soup = BeautifulSoup(response.text, "html.parser")
        # 将源码保存到文件
        with open(save_path, "w", encoding="utf-8") as file:
            file.write(soup.prettify())


def setup_logging():
    # 获取当前时间，并格式化为字符串
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    logfilename = f"D:\\football\\stats\\log\\footballstats_{current_time}.log"

    # 配置日志
    logging.basicConfig(
        filename=logfilename,
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )


def run(year, xlsx_path):
    setup_logging()
    # 尝试加载现有的Excel文件，如果不存在则创建一个新的
    try:
        wb = load_workbook(xlsx_path)
    except FileNotFoundError:
        wb = Workbook()

    sheet_name = f"{year}-{year+1}"
    # 检查工作表是否已存在，如果存在则删除
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    match_reports = get_match_report(
        "D:\\football\\stats\\html\\B_goal\\B_goal_stats_"
        + str(year)
        + "-"
        + str(year + 1)
        + ".html"
    )
    logging.info(f"Found {len(match_reports)} match reports.")

    # 写入表头
    ws.append(["Report", "Goal Time (mins)"])

    for i, report in enumerate(match_reports):
        try:
            time.sleep(random.uniform(3, 6))  # 随机暂停以减少被限速的风险
            headers = get_headers()
            times = get_goal_times("https://fbref.com/" + report, headers=headers)
            if times:  # 如果times不为空，则添加到工作簿中
                for t in times:
                    ws.append([report, t])
                logging.info(
                    f"Processed match report {i+1}/{len(match_reports)}: {report}"
                )
                print(f"Processed match report {i+1}/{len(match_reports)}: {report}")
                logging.info(f"Goal times added: {times}")
                print(f"Goal times added: {times}")
        except Exception as e:
            logging.error(f"Error processing match report {i+1}: {report}. Error: {e}")
            print(f"Error processing match report {i+1}: {report}. Error: {e}")

    # 保存工作簿
    wb.save(xlsx_path)
    logging.info(f"Data saved to Excel file: {xlsx_path}")
    print(f"Data saved to Excel file: {xlsx_path}")


# 英超--https://fbref.com/en/comps/9/2023-2024/schedule/2023-2024-Premier-League-Scores-and-Fixtures
# 意甲--https://fbref.com/en/comps/11/2023-2024/schedule/2023-2024-Serie-A-Scores-and-Fixtures
# 西甲--https://fbref.com/en/comps/12/2023-2024/schedule/2023-2024-La-Liga-Scores-and-Fixtures
# 法甲--https://fbref.com/en/comps/13/2023-2024/schedule/2023-2024-Ligue-1-Scores-and-Fixtures
# 德甲--https://fbref.com/en/comps/20/2023-2024/schedule/2023-2024-Bundesliga-Scores-and-Fixtures

for year in range(2004, 2024):
    get_html(
        "https://fbref.com/en/comps/20/"
        + str(year)
        + "-"
        + str(year + 1)
        + "/schedule/",
        "D:\\football\\stats\\html\\B_goal\\B_goal_stats_"
        + str(year)
        + "-"
        + str(year + 1)
        + ".html",
    )
    time.sleep(5)

for i in range(2004, 2024):
    run(i, "D:\\football\\stats\\data\\B_goal\\B_goals.xlsx")
